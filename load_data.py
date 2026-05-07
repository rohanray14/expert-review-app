import json
import openpyxl
from pathlib import Path

BASE = Path(__file__).resolve().parent / "data"


def _safe_json(val):
    if not val:
        return []
    try:
        return json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return []


def load_posts():
    """Return dict keyed by (post_id, model_name) with all LLM output fields."""
    wb = openpyxl.load_workbook(BASE / "PostLevel_Outputs.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    posts = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        key = (d["post_id"], d["model_name"])
        posts[key] = {
            "class_label": d["class_label"],
            "post_id": d["post_id"],
            "title": d["title"],
            "link": d["link"],
            "model_family": d["model_family"],
            "model_name": d["model_name"],
            "summary": d["summary"] or "",
            "advice": _safe_json(d["unique_advice_json"]),
            "divergences": _safe_json(d["divergences_json"]),
            "clinical_notes": _safe_json(d["clinically_relevant_notes_json"]),
            "data_quality": d["data_quality"] or "",
        }
    wb.close()
    return posts


def load_comments():
    """Return dict keyed by post_id with title, body, and list of comments."""
    wb = openpyxl.load_workbook(BASE / "6K_data_with_comments.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    comment_cols = [h for h in headers if h and h.startswith("Comment")]
    data = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        pid = d["id"]
        comments = []
        for col in comment_cols:
            val = d.get(col)
            if val and str(val).strip():
                comments.append(str(val).strip())
        data[pid] = {
            "post_id": pid,
            "title": d.get("title", ""),
            "body": d.get("body", ""),
            "label1": d.get("Label1", ""),
            "label2": d.get("Label2", ""),
            "label3": d.get("Label3", ""),
            "comments": comments,
        }
    wb.close()
    return data


def load_all():
    """Load and merge posts with their comments. Returns list of unique post_ids and full data."""
    posts = load_posts()
    comments = load_comments()

    # Get unique post_ids preserving order
    seen = set()
    post_ids = []
    for (pid, _) in posts:
        if pid not in seen:
            seen.add(pid)
            post_ids.append(pid)

    # Get model names
    models = sorted(set(m for (_, m) in posts))

    return post_ids, posts, comments, models
